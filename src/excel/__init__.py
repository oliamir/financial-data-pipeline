"""Excel financial model builder.

Generates a professional 9-sheet Excel workbook from financial data
with formatted tables, charts, and dashboards.

Color-coding convention:
    - Blue font (#0000FF): historical data inputs
    - Black font (#000000): formulas / calculated cells
    - Green font (#008000): cross-sheet references
    - Yellow fill: key assumption cells
"""

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from ..models.financial import FinancialPeriod
from ..models.kpi import KPIMetrics
from ..storage.file_manager import FileManager
from ..utils.logging import get_logger

logger = get_logger(__name__)

# Sheet configuration
SHEETS = [
    "Dashboard",
    "Income Statement",
    "Balance Sheet",
    "Cash Flow",
    "KPI Ratios",
    "Growth Analysis",
    "Valuation",
    "Peer Comparison",
    "Notes",
]


class ExcelModelBuilder:
    """Builds a professional financial model Excel workbook."""

    def __init__(self, slug: str):
        self.slug = slug
        self.storage = FileManager(slug)

    def build(self, output_path: Optional[str] = None) -> str:
        """Build the complete Excel model.

        Args:
            output_path: Override output file path.

        Returns:
            Path to the generated Excel file.
        """
        if output_path is None:
            output_path = str(self.storage.paths.model_xlsx)

        # Load data
        periods = self.storage.load_financials()
        kpi = self.storage.load_kpis()

        if not periods:
            logger.warning(f"No financial data for {self.slug}, generating empty model")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, Reference, LineChart
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # Create sheets
            for i, sheet_name in enumerate(SHEETS):
                if i == 0:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    wb.create_sheet(sheet_name)

            # Populate sheets
            self._build_dashboard(wb["Dashboard"], periods, kpi)
            self._build_income_statement(wb["Income Statement"], periods)
            self._build_balance_sheet(wb["Balance Sheet"], periods)
            self._build_cash_flow(wb["Cash Flow"], periods)
            self._build_kpi_ratios(wb["KPI Ratios"], periods, kpi)
            self._build_growth_analysis(wb["Growth Analysis"], periods)
            self._build_valuation(wb["Valuation"], periods, kpi)
            self._build_peer_comparison(wb["Peer Comparison"])
            self._build_notes(wb["Notes"])

            # Charts (added to Dashboard after all data sheets are populated)
            self._add_dashboard_charts(wb["Dashboard"], periods)

            # Save
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            logger.info(f"Excel model saved: {output_path}")
            return output_path

        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise

    # ------------------------------------------------------------------
    # Style helpers
    # ------------------------------------------------------------------

    def _header_style(self) -> Dict[str, Any]:
        """Return style kwargs for column-header cells."""
        from openpyxl.styles import Font, PatternFill, Alignment

        return {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center"),
        }

    def _number_format(self, value: Any, is_pct: bool = False) -> str:
        """Pick the right number format string for *value*."""
        if is_pct:
            return "0.0%"
        if isinstance(value, float) and abs(value) < 100:
            return "#,##0.00"
        return "#,##0"

    @staticmethod
    def _font_historical():
        """Blue font for historical data inputs."""
        from openpyxl.styles import Font
        return Font(color="0000FF")

    @staticmethod
    def _font_calculated():
        """Black font for formulas / calculated cells."""
        from openpyxl.styles import Font
        return Font(color="000000")

    @staticmethod
    def _font_crossref():
        """Green font for cross-sheet references."""
        from openpyxl.styles import Font
        return Font(color="008000")

    @staticmethod
    def _fill_assumption():
        """Yellow fill for key assumption cells."""
        from openpyxl.styles import PatternFill
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    @staticmethod
    def _fmt_negative_parens() -> str:
        """Number format that shows negatives in parentheses."""
        return '#,##0_);(#,##0)'

    @staticmethod
    def _fmt_negative_parens_dec() -> str:
        """Decimal number format with negatives in parentheses."""
        return '#,##0.00_);(#,##0.00)'

    def _apply_sheet_defaults(self, ws, periods: List[FinancialPeriod]) -> None:
        """Apply common formatting to a financial-statement sheet.

        Sets column widths, freezes the header row + label column, and
        widens column A for line-item labels.
        """
        from openpyxl.utils import get_column_letter

        ws.column_dimensions["A"].width = 38
        for col_idx in range(2, 2 + len(periods)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        ws.freeze_panes = "B2"

    def _build_dashboard(self, ws, periods: List[FinancialPeriod], kpi: Optional[KPIMetrics]):
        """Build the Dashboard summary sheet."""
        from openpyxl.styles import Font as _F

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 18

        # Title
        ws["A1"] = f"Financial Dashboard: {self.slug.upper()}"
        ws["A1"].font = _F(bold=True, size=16, color="1F4E79")

        row = 3
        if periods:
            latest = periods[-1]
            ws[f"A{row}"] = "Latest Period"
            ws[f"B{row}"] = f"{latest.period_type} {latest.fiscal_year}"
            row += 1
            ws[f"A{row}"] = "Currency"
            ws[f"B{row}"] = latest.currency
            row += 1

            if latest.income_statement.revenue:
                ws[f"A{row}"] = "Revenue"
                ws[f"B{row}"] = latest.income_statement.revenue
                ws[f"B{row}"].number_format = self._fmt_negative_parens()
                ws[f"B{row}"].font = self._font_crossref()
                row += 1

            if latest.income_statement.net_income:
                ws[f"A{row}"] = "Net Income"
                ws[f"B{row}"] = latest.income_statement.net_income
                ws[f"B{row}"].number_format = self._fmt_negative_parens()
                ws[f"B{row}"].font = self._font_crossref()
                row += 1

        if kpi:
            row += 1
            ws[f"A{row}"] = "Key Ratios"
            ws[f"A{row}"].font = _F(bold=True, size=12, color="1F4E79")
            row += 1
            for label, value in [
                ("Gross Margin", kpi.gross_margin),
                ("Operating Margin", kpi.operating_margin),
                ("Net Margin", kpi.net_margin),
                ("Revenue Growth", kpi.revenue_growth),
                ("ROE", kpi.roe),
                ("Debt/Equity", kpi.debt_to_equity),
            ]:
                if value is not None:
                    ws[f"A{row}"] = label
                    ws[f"B{row}"] = value
                    ws[f"B{row}"].number_format = "0.0%"
                    ws[f"B{row}"].font = self._font_crossref()
                    row += 1

        # Store the last used row so chart placement knows where to start
        self._dashboard_data_end_row = row

    def _build_income_statement(self, ws, periods: List[FinancialPeriod]):
        """Build Income Statement sheet with multi-period columns."""
        headers = ["Line Item"] + [f"{p.period_type} {p.fiscal_year}" for p in periods]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            for k, v in self._header_style().items():
                setattr(cell, k, v)

        self._apply_sheet_defaults(ws, periods)

        # Fields: (label, attribute, is_calculated)
        # Calculated items (subtotals / derived) get black font; raw inputs get blue.
        fields = [
            ("Revenue", "revenue", False),
            ("Cost of Revenue", "cost_of_revenue", False),
            ("Gross Profit", "gross_profit", True),
            ("R&D Expense", "rd_expense", False),
            ("SG&A Expense", "sga_expense", False),
            ("Operating Income", "operating_income", True),
            ("Interest Expense", "interest_expense", False),
            ("Pretax Income", "pretax_income", True),
            ("Income Tax", "income_tax", False),
            ("Net Income", "net_income", True),
            ("EBITDA", "ebitda", True),
            ("EPS (Diluted)", "eps_diluted", True),
        ]

        for row_idx, (label, field, calculated) in enumerate(fields, 2):
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            if calculated:
                from openpyxl.styles import Font as _F
                label_cell.font = _F(bold=True)
            for col_idx, period in enumerate(periods, 2):
                val = getattr(period.income_statement, field, None)
                if val is not None:
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = self._font_calculated() if calculated else self._font_historical()
                    if field == "eps_diluted":
                        cell.number_format = self._fmt_negative_parens_dec()
                    else:
                        cell.number_format = self._fmt_negative_parens()

    def _build_balance_sheet(self, ws, periods: List[FinancialPeriod]):
        """Build Balance Sheet sheet."""
        headers = ["Line Item"] + [f"{p.period_type} {p.fiscal_year}" for p in periods]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            for k, v in self._header_style().items():
                setattr(cell, k, v)

        self._apply_sheet_defaults(ws, periods)

        # (label, attribute, is_subtotal)
        fields = [
            ("Cash & Equivalents", "cash_and_equivalents", False),
            ("Accounts Receivable", "accounts_receivable", False),
            ("Inventory", "inventory", False),
            ("Total Current Assets", "total_current_assets", True),
            ("PP&E Net", "ppe_net", False),
            ("Total Assets", "total_assets", True),
            ("Accounts Payable", "accounts_payable", False),
            ("Total Current Liabilities", "total_current_liabilities", True),
            ("Long-term Debt", "long_term_debt", False),
            ("Total Liabilities", "total_liabilities", True),
            ("Total Equity", "total_equity", True),
        ]

        for row_idx, (label, field, subtotal) in enumerate(fields, 2):
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            if subtotal:
                from openpyxl.styles import Font as _F
                label_cell.font = _F(bold=True)
            for col_idx, period in enumerate(periods, 2):
                val = getattr(period.balance_sheet, field, None) if period.balance_sheet else None
                if val is not None:
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = self._font_calculated() if subtotal else self._font_historical()
                    cell.number_format = self._fmt_negative_parens()

    def _build_cash_flow(self, ws, periods: List[FinancialPeriod]):
        """Build Cash Flow sheet."""
        headers = ["Line Item"] + [f"{p.period_type} {p.fiscal_year}" for p in periods]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            for k, v in self._header_style().items():
                setattr(cell, k, v)

        self._apply_sheet_defaults(ws, periods)

        # (label, attribute, is_subtotal)
        fields = [
            ("Cash from Operations", "cash_from_operations", True),
            ("CapEx", "capex", False),
            ("Free Cash Flow", "free_cash_flow", True),
            ("Cash from Investing", "cash_from_investing", True),
            ("Cash from Financing", "cash_from_financing", True),
            ("Net Change in Cash", "net_change_in_cash", True),
        ]

        for row_idx, (label, field, subtotal) in enumerate(fields, 2):
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            if subtotal:
                from openpyxl.styles import Font as _F
                label_cell.font = _F(bold=True)
            for col_idx, period in enumerate(periods, 2):
                val = getattr(period.cash_flow, field, None) if period.cash_flow else None
                if val is not None:
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = self._font_calculated() if subtotal else self._font_historical()
                    cell.number_format = self._fmt_negative_parens()

    def _build_kpi_ratios(self, ws, periods: List[FinancialPeriod], kpi: Optional[KPIMetrics]):
        """Build KPI Ratios sheet.

        Values here are derived from Income Statement / Balance Sheet data,
        so they use the green cross-sheet-reference font.
        """
        from openpyxl.styles import Font as _F

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 15
        ws.freeze_panes = "A2"

        for col, h in enumerate(["KPI Ratio", "Value"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            for k, v in self._header_style().items():
                setattr(cell, k, v)

        if not kpi:
            ws.cell(row=2, column=1, value="No KPI data available")
            return

        ratios = [
            ("Gross Margin", kpi.gross_margin),
            ("Operating Margin", kpi.operating_margin),
            ("EBITDA Margin", kpi.ebitda_margin),
            ("Net Margin", kpi.net_margin),
            ("R&D % Revenue", kpi.rd_pct_revenue),
            ("SG&A % Revenue", kpi.sga_pct_revenue),
            ("Revenue Growth", kpi.revenue_growth),
            ("ROE", kpi.roe),
            ("ROA", kpi.roa),
            ("Debt/Equity", kpi.debt_to_equity),
            ("Current Ratio", kpi.current_ratio),
            ("Interest Coverage", kpi.interest_coverage),
            ("P/E Ratio", kpi.pe_ratio),
            ("EV/EBITDA", kpi.ev_to_ebitda),
        ]

        for row_idx, (label, value) in enumerate(ratios, 2):
            ws.cell(row=row_idx, column=1, value=label)
            if value is not None:
                cell = ws.cell(row=row_idx, column=2, value=value)
                cell.font = self._font_crossref()
                if any(kw in label for kw in ["Margin", "Growth", "ROE", "ROA", "%"]):
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "#,##0.00"

    def _build_growth_analysis(self, ws, periods: List[FinancialPeriod]):
        """Build Growth Analysis sheet with period-over-period changes."""
        from openpyxl.styles import Font as _F
        from openpyxl.utils import get_column_letter

        ws.column_dimensions["A"].width = 38

        ws.cell(row=1, column=1, value="Growth Analysis").font = _F(bold=True, size=14, color="1F4E79")

        if len(periods) < 2:
            ws.cell(row=3, column=1, value="Need at least 2 periods for growth analysis")
            return

        headers = ["Metric"] + [f"{p.period_type} {p.fiscal_year}" for p in periods[1:]]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            for k, v in self._header_style().items():
                setattr(cell, k, v)
            if col >= 2:
                ws.column_dimensions[get_column_letter(col)].width = 15

        ws.freeze_panes = "B4"

        metrics = ["revenue", "gross_profit", "operating_income", "net_income", "ebitda"]
        labels = ["Revenue", "Gross Profit", "Operating Income", "Net Income", "EBITDA"]

        for row_idx, (label, field) in enumerate(zip(labels, metrics), 4):
            ws.cell(row=row_idx, column=1, value=f"{label} Growth")
            for col_idx, i in enumerate(range(1, len(periods)), 2):
                curr_val = getattr(periods[i].income_statement, field, None)
                prev_val = getattr(periods[i - 1].income_statement, field, None)
                if curr_val and prev_val and prev_val != 0:
                    growth = (curr_val - prev_val) / abs(prev_val)
                    cell = ws.cell(row=row_idx, column=col_idx, value=growth)
                    cell.number_format = "0.0%"
                    cell.font = self._font_calculated()

    def _build_valuation(self, ws, periods: List[FinancialPeriod], kpi: Optional[KPIMetrics]):
        """Build Valuation sheet.

        Valuation multiples depend on market-price assumptions, so the value
        cells get a yellow fill to flag them as assumption-sensitive.
        """
        from openpyxl.styles import Font as _F

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 15
        ws.freeze_panes = "A3"

        ws.cell(row=1, column=1, value="Valuation Metrics").font = _F(bold=True, size=14, color="1F4E79")

        if not kpi:
            ws.cell(row=3, column=1, value="No valuation data available")
            return

        row = 3
        for label, value in [
            ("P/E Ratio", kpi.pe_ratio),
            ("EV/EBITDA", kpi.ev_to_ebitda),
            ("Price/Book", kpi.price_to_book),
            ("Price/Sales", kpi.price_to_sales),
            ("Price/FCF", kpi.price_to_fcf),
            ("Dividend Yield", kpi.dividend_yield),
        ]:
            ws.cell(row=row, column=1, value=label)
            if value is not None:
                cell = ws.cell(row=row, column=2, value=value)
                cell.number_format = "#,##0.00" if "Yield" not in label else "0.0%"
                cell.font = self._font_crossref()
                cell.fill = self._fill_assumption()
            row += 1

    # ------------------------------------------------------------------
    # Charts
    # ------------------------------------------------------------------

    def _add_dashboard_charts(self, ws, periods: List[FinancialPeriod]) -> None:
        """Add revenue trend (bar) and margin evolution (line) charts to Dashboard.

        Charts are embedded in the Dashboard sheet below the summary KPIs.
        A small helper data table is written in columns D-G (out of the way)
        so that openpyxl ``Reference`` objects can point at real worksheet
        cells.
        """
        if len(periods) < 2:
            return  # Not enough data to chart

        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.utils import get_column_letter

        # ---- Write helper data table in columns D onward ----
        # Row layout (starting at row 1):
        #   Row 1: period labels
        #   Row 2: revenue
        #   Row 3: gross margin (decimal)
        #   Row 4: operating margin (decimal)
        data_start_col = 4  # column D
        header_row = 1
        rev_row = 2
        gm_row = 3
        om_row = 4

        ws.cell(row=header_row, column=data_start_col, value="Period")
        ws.cell(row=rev_row, column=data_start_col, value="Revenue")
        ws.cell(row=gm_row, column=data_start_col, value="Gross Margin")
        ws.cell(row=om_row, column=data_start_col, value="Oper. Margin")

        for idx, period in enumerate(periods):
            col = data_start_col + 1 + idx
            label = f"{period.period_type} {period.fiscal_year}"
            ws.cell(row=header_row, column=col, value=label)

            rev = period.income_statement.revenue
            ws.cell(row=rev_row, column=col, value=rev if rev is not None else 0)

            # Compute margins inline (safe division)
            gm = None
            om = None
            if rev and rev != 0:
                gp = period.income_statement.gross_profit
                oi = period.income_statement.operating_income
                if gp is not None:
                    gm = gp / rev
                if oi is not None:
                    om = oi / rev
            ws.cell(row=gm_row, column=col, value=gm if gm is not None else 0)
            ws.cell(row=om_row, column=col, value=om if om is not None else 0)

        last_data_col = data_start_col + len(periods)

        # ---- Revenue trend bar chart ----
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Revenue Trend"
        bar.y_axis.title = "Revenue"
        bar.x_axis.title = "Period"
        bar.y_axis.numFmt = "#,##0"

        cats = Reference(ws, min_col=data_start_col + 1, max_col=last_data_col, min_row=header_row)
        rev_data = Reference(ws, min_col=data_start_col + 1, max_col=last_data_col, min_row=rev_row)
        bar.add_data(rev_data, from_rows=True, titles_from_data=False)
        bar.set_categories(cats)
        bar.series[0].title = "Revenue"
        bar.width = 18
        bar.height = 12

        chart_start_row = getattr(self, "_dashboard_data_end_row", 18) + 2
        ws.add_chart(bar, f"A{chart_start_row}")

        # ---- Margin evolution line chart ----
        line = LineChart()
        line.style = 10
        line.title = "Margin Evolution"
        line.y_axis.title = "Margin %"
        line.x_axis.title = "Period"
        line.y_axis.numFmt = "0.0%"

        cats_line = Reference(ws, min_col=data_start_col + 1, max_col=last_data_col, min_row=header_row)
        gm_data = Reference(ws, min_col=data_start_col + 1, max_col=last_data_col, min_row=gm_row)
        om_data = Reference(ws, min_col=data_start_col + 1, max_col=last_data_col, min_row=om_row)

        line.add_data(gm_data, from_rows=True, titles_from_data=False)
        line.add_data(om_data, from_rows=True, titles_from_data=False)
        line.set_categories(cats_line)

        line.series[0].title = "Gross Margin"
        line.series[1].title = "Operating Margin"

        line.width = 18
        line.height = 12

        ws.add_chart(line, f"A{chart_start_row + 16}")

    def _build_peer_comparison(self, ws):
        """Build Peer Comparison sheet with a placeholder comp table.

        Generates a structured table that can be filled with comparable
        company data either manually or by a downstream data loader.
        """
        from openpyxl.styles import Font as _F, Border, Side

        ws.column_dimensions["A"].width = 28
        for col_letter in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 16

        ws.cell(row=1, column=1, value="Peer Comparison").font = _F(bold=True, size=14, color="1F4E79")
        ws.cell(row=2, column=1, value="(populate with comparable companies)").font = _F(italic=True, color="808080")

        # Header row
        peer_headers = [
            "Company",
            "Market Cap",
            "Revenue",
            "EV/EBITDA",
            "P/E Ratio",
            "Gross Margin",
        ]
        for col, h in enumerate(peer_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            for k, v in self._header_style().items():
                setattr(cell, k, v)

        ws.freeze_panes = "A5"

        # Pre-fill the subject company row
        ws.cell(row=5, column=1, value=self.slug.upper())
        ws.cell(row=5, column=1).font = _F(bold=True, color="0000FF")

        # Placeholder peer rows
        thin_border = Border(
            bottom=Side(style="thin", color="D9D9D9"),
        )
        for peer_row in range(6, 11):
            ws.cell(row=peer_row, column=1, value=f"Peer {peer_row - 5}")
            ws.cell(row=peer_row, column=1).font = _F(color="808080", italic=True)
            for col in range(1, len(peer_headers) + 1):
                ws.cell(row=peer_row, column=col).border = thin_border

        # Format hints for data columns
        for data_row in range(5, 11):
            ws.cell(row=data_row, column=2).number_format = "#,##0"       # Market Cap
            ws.cell(row=data_row, column=3).number_format = "#,##0"       # Revenue
            ws.cell(row=data_row, column=4).number_format = "#,##0.0x"    # EV/EBITDA
            ws.cell(row=data_row, column=5).number_format = "#,##0.0x"    # P/E
            ws.cell(row=data_row, column=6).number_format = "0.0%"        # Gross Margin

        # Summary stats row
        summary_row = 12
        ws.cell(row=summary_row, column=1, value="Peer Median").font = _F(bold=True)
        ws.cell(row=summary_row + 1, column=1, value="Peer Mean").font = _F(bold=True)
        ws.cell(row=summary_row + 2, column=1, value="Premium / (Discount)").font = _F(bold=True)
        for stat_row in range(summary_row, summary_row + 3):
            for col in range(2, len(peer_headers) + 1):
                ws.cell(row=stat_row, column=col).font = self._font_calculated()

    def _build_notes(self, ws):
        """Build Notes sheet with metadata."""
        from datetime import datetime
        from openpyxl.styles import Font as _F

        ws.column_dimensions["A"].width = 60
        ws.cell(row=1, column=1, value="Model Notes").font = _F(bold=True, size=14, color="1F4E79")

        ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=4, column=1, value=f"Company: {self.slug}")
        ws.cell(row=5, column=1, value="Source: Financial Data Pipeline v3")

        # Color-coding legend
        ws.cell(row=7, column=1, value="Color-Coding Legend").font = _F(bold=True, size=12)
        ws.cell(row=8, column=1, value="Blue font = Historical data inputs")
        ws.cell(row=8, column=1).font = self._font_historical()
        ws.cell(row=9, column=1, value="Black font = Formulas / calculated cells")
        ws.cell(row=9, column=1).font = self._font_calculated()
        ws.cell(row=10, column=1, value="Green font = Cross-sheet references")
        ws.cell(row=10, column=1).font = self._font_crossref()
        ws.cell(row=11, column=1, value="Yellow fill = Key assumption cells")
        ws.cell(row=11, column=1).fill = self._fill_assumption()

        ws.cell(row=13, column=1, value="Disclaimer: This model is auto-generated from parsed financial reports.")
        ws.cell(row=14, column=1, value="All data should be verified against original source documents.")
